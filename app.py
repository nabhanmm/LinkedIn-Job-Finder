"""
LinkedIn Job Finder — v4
Synced with linkedin_jobs.py v6:
  - primary_must_have / secondary_must_have / Score columns
  - Skills filter removed (replaced by keyword scoring)
  - Match column replaced by Score
  - Hover tooltips on every filter (info icon)
  - Font / cursor / visibility fixes
"""

import io
from datetime import datetime

import pandas as pd
import streamlit as st
from supabase import create_client

import linkedin_jobs as lj

# ─────────────────────────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="LinkedIn Job Finder",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────
# CSS — fixes all visibility issues + hover tooltips
# ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600&family=DM+Mono&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }

/* ── Sidebar background ── */
[data-testid="stSidebar"] { background: #0f172a; }
[data-testid="stSidebar"] * { color: #e2e8f0 !important; }

/* ── Sidebar inputs — dark background, DARK text so it's visible ── */
[data-testid="stSidebar"] .stTextInput input,
[data-testid="stSidebar"] .stTextArea textarea {
    background: #1e293b !important;
    border: 1px solid #475569 !important;
    color: #f1f5f9 !important;
    caret-color: #ef4444 !important;
}
[data-testid="stSidebar"] .stTextInput input::placeholder,
[data-testid="stSidebar"] .stTextArea textarea::placeholder {
    color: #64748b !important;
}

/* ── Selectbox — make selected value text dark/visible ── */
[data-testid="stSidebar"] .stSelectbox [data-baseweb="select"] > div,
[data-testid="stSidebar"] .stSelectbox [data-baseweb="select"] span,
[data-testid="stSidebar"] .stSelectbox [data-baseweb="select"] div[class*="ValueContainer"] {
    background: #1e293b !important;
    color: #f1f5f9 !important;
}
[data-testid="stSidebar"] .stSelectbox [data-baseweb="select"] * {
    color: #f1f5f9 !important;
}
/* Fix Sign Out button text */
[data-testid="stSidebar"] .stButton > button {
    color: #1e293b !important;
    background: #e2e8f0 !important;
    border: none !important;
    border-radius: 6px !important;
    font-weight: 500 !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: #cbd5e1 !important;
}

/* ── Multiselect tags ── */
[data-testid="stSidebar"] .stMultiSelect > div {
    background: #1e293b !important;
    border: 1px solid #475569 !important;
}
[data-testid="stSidebar"] .stMultiSelect span[data-baseweb="tag"] {
    background: #2563eb !important;
}

/* ── Select slider ── */
[data-testid="stSidebar"] .stSlider * { color: #e2e8f0 !important; }

/* ── Number input ── */
[data-testid="stSidebar"] .stNumberInput input {
    background: #1e293b !important;
    border: 1px solid #475569 !important;
    color: #f1f5f9 !important;
}

/* ── Metrics ── */
[data-testid="metric-container"] {
    background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 10px; padding: 12px 16px;
}
[data-testid="stMetricValue"] { font-size: 2rem !important; font-weight: 600 !important; color: #0f172a !important; }
[data-testid="stMetricLabel"] { color: #64748b !important; font-size: 0.8rem !important; text-transform: uppercase; letter-spacing: 0.05em; }

/* ── Primary action button ── */
.stButton > button[kind="primary"] {
    background: #2563eb !important; color: white !important; border: none !important;
    border-radius: 8px !important; font-weight: 600 !important; height: 48px !important; font-size: 1rem !important;
}
.stButton > button[kind="primary"]:hover { background: #1d4ed8 !important; }

/* ── Download buttons ── */
.stDownloadButton > button { border-radius: 8px !important; font-weight: 500 !important; }

/* ── Dataframe ── */
[data-testid="stDataFrame"] { border: 1px solid #e2e8f0; border-radius: 10px; overflow: hidden; }

/* ── Live log box ── */
.log-box {
    background: #0f172a; color: #a3e635; font-family: 'DM Mono', monospace;
    font-size: 0.78rem; padding: 12px 16px; border-radius: 8px;
    max-height: 200px; overflow-y: auto; white-space: pre-wrap; line-height: 1.6;
}

hr { border-color: #e2e8f0 !important; margin: 1.2rem 0 !important; }

/* ── Tooltip (ℹ hover) ── */
.filter-label-row {
    display: flex;
    align-items: center;
    gap: 6px;
    margin-bottom: 4px;
    color: #e2e8f0;
    font-size: 13px;
}
.info-icon {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 16px; height: 16px;
    border-radius: 50%;
    background: #334155;
    color: #94a3b8 !important;
    font-size: 10px;
    font-weight: 700;
    cursor: help;
    position: relative;
    flex-shrink: 0;
    line-height: 1;
}
.info-icon:hover { background: #2563eb; color: #fff !important; }
.info-icon .tooltip-text {
    visibility: hidden;
    opacity: 0;
    width: 220px;
    background: #1e293b;
    color: #e2e8f0;
    font-size: 11px;
    line-height: 1.5;
    text-align: left;
    padding: 8px 10px;
    border-radius: 7px;
    border: 1px solid #334155;
    position: absolute;
    left: 22px;
    top: -4px;
    z-index: 9999;
    transition: opacity 0.15s;
    font-weight: 400;
    pointer-events: none;
}
.info-icon:hover .tooltip-text {
    visibility: visible;
    opacity: 1;
}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────
# Tooltip helper — renders label + ℹ icon with hover tooltip
# ─────────────────────────────────────────────────────────────────
def filter_label(icon, label, tip):
    st.markdown(f"""
    <div class="filter-label-row">
      <span>{icon} {label}</span>
      <span class="info-icon">i
        <span class="tooltip-text">{tip}</span>
      </span>
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────
# Supabase client
# ─────────────────────────────────────────────────────────────────
@st.cache_resource
def get_supabase():
    return create_client(
        st.secrets["SUPABASE_URL"],
        st.secrets["SUPABASE_ANON_KEY"],
    )

supabase = get_supabase()


# ─────────────────────────────────────────────────────────────────
# Company size → sortable prefix so column sorts correctly
# ─────────────────────────────────────────────────────────────────
def sortable_company_size(s):
    """Return a prefix-letter version (A-H) that sorts correctly ascending/descending."""
    if not s:
        return ""
    digits = ""
    for ch in str(s):
        if ch.isdigit():
            digits += ch
        elif ch == "," and digits:
            continue          # skip commas inside numbers e.g. 10,001
        elif digits:
            break             # stop at first non-digit/non-comma after digits start
    if not digits:
        return str(s)
    n = int(digits)
    if n >= 10001: return "H: 10,001+"
    if n >= 5001:  return "G: 5,001-10,000"
    if n >= 1001:  return "F: 1,001-5,000"
    if n >= 501:   return "E: 501-1,000"
    if n >= 201:   return "D: 201-500"
    if n >= 51:    return "C: 51-200"
    if n >= 11:    return "B: 11-50"
    return                "A: 1-10"


# ─────────────────────────────────────────────────────────────────
# User preferences — save & load last-used filters
# ─────────────────────────────────────────────────────────────────
def load_user_prefs(user_id: str) -> dict:
    try:
        res = (supabase.table("user_preferences")
               .select("preferences")
               .eq("user_id", user_id)
               .execute())
        if res.data:
            return res.data[0].get("preferences") or {}
    except Exception:
        pass
    return {}

def save_user_prefs(user_id: str, prefs: dict):
    try:
        supabase.table("user_preferences").upsert({
            "user_id":      user_id,
            "preferences":  prefs,
            "updated_at":   datetime.now().isoformat(),
        }).execute()
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────────
# Session state
# ─────────────────────────────────────────────────────────────────
DEFAULTS = {
    "user": None, "access_token": None, "refresh_token": None,
    "results": None, "new_count": 0, "dup_count": 0,
}
for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

if st.session_state.access_token and st.session_state.user is None:
    try:
        # Restore + silently refresh the token so the session extends
        res = supabase.auth.set_session(st.session_state.access_token, st.session_state.refresh_token)
        try:
            refreshed = supabase.auth.refresh_session()
            st.session_state.access_token  = refreshed.session.access_token
            st.session_state.refresh_token = refreshed.session.refresh_token
            st.session_state.user          = refreshed.user
        except Exception:
            st.session_state.user = res.user

        # Load saved filter preferences into session state (only for fresh sessions)
        if not st.session_state.get("prefs_loaded") and st.session_state.user:
            prefs = load_user_prefs(st.session_state.user.id)
            for k, v in prefs.items():
                if k not in st.session_state:
                    st.session_state[k] = v
            st.session_state.prefs_loaded = True

    except Exception:
        st.session_state.access_token = st.session_state.refresh_token = None


# ─────────────────────────────────────────────────────────────────
# Auth helpers
# ─────────────────────────────────────────────────────────────────
def do_signup(email, password, username):
    try:
        res = supabase.auth.sign_up({
            "email": email, "password": password,
            "options": {"data": {"username": username}},
        })
        if res.user is None:
            return None, "Sign up failed. The email may already be registered."
        return res.user, None
    except Exception as e:
        err = str(e)
        if "already registered" in err.lower() or "already exists" in err.lower():
            return None, "This email is already registered. Please Sign In instead."
        return None, err

def do_signin(email, password):
    try:
        res = supabase.auth.sign_in_with_password({"email": email, "password": password})
        return res, None
    except Exception as e:
        err = str(e)
        if "invalid" in err.lower() or "credentials" in err.lower():
            return None, "Invalid email or password. Please check and try again."
        if "confirm" in err.lower():
            return None, "Please verify your email first — check your inbox."
        return None, err

def do_signout():
    try:
        supabase.auth.sign_out()
    except Exception:
        pass
    for k in ["user", "access_token", "refresh_token", "results"]:
        st.session_state[k] = None
    st.rerun()


# ─────────────────────────────────────────────────────────────────
# Repository DB helpers
# ─────────────────────────────────────────────────────────────────
def get_repository(user_id):
    res = (supabase.table("job_repository").select("*")
           .eq("user_id", user_id).order("date_added", desc=True).execute())
    return res.data or []

def get_existing_urls(user_id):
    res = (supabase.table("job_repository").select("linkedin_url")
           .eq("user_id", user_id).execute())
    return {r["linkedin_url"] for r in (res.data or []) if r.get("linkedin_url")}

def append_new_jobs(jobs, user):
    existing  = get_existing_urls(user.id)
    new_jobs  = [j for j in jobs if j.get("LinkedIn URL", "") not in existing]
    dup_count = len(jobs) - len(new_jobs)
    if not new_jobs:
        return 0, dup_count
    username = (user.user_metadata or {}).get("username", "")
    rows = [{
        "user_id":                   user.id,
        "user_email":                user.email,
        "username":                  username,
        "job_title":                 j.get("Job Title", ""),
        "company":                   j.get("Company", ""),
        "company_size":              j.get("Company Size", ""),
        "company_type":              j.get("Company Type", ""),
        "level":                     j.get("Level", ""),
        "location":                  j.get("Location", ""),
        "work_mode":                 j.get("Work Mode", ""),
        "linkedin_url":              j.get("LinkedIn URL", ""),
        "score":                     str(j.get("Score", "")),
        "primary_keywords_match":    j.get("Primary Keywords Match", ""),
        "secondary_keywords_match":  j.get("Secondary Keywords Match", ""),
        "easy_apply":                j.get("Easy Apply", ""),
        "still_accepting":           j.get("Still Accepting?", ""),
        "posted_date":               j.get("Posted Date", ""),
        "date_searched":             j.get("Date Searched", ""),
        "exp_required":              j.get("Exp. Required", ""),
        "recruiter_profile":         j.get("Recruiter Profile", ""),
        "status":                    "New",
        "date_added":                datetime.now().isoformat(),
    } for j in new_jobs]
    supabase.table("job_repository").insert(rows).execute()
    return len(new_jobs), dup_count

def update_job_status(row_id, status):
    supabase.table("job_repository").update({"status": status}).eq("id", row_id).execute()


# ─────────────────────────────────────────────────────────────────
# AUTH PAGE
# ─────────────────────────────────────────────────────────────────
def show_auth_page():
    st.markdown("""
    <div style='text-align:center; padding:40px 0 24px;'>
        <div style='font-size:3.5rem;'>🔍</div>
        <h1 style='font-size:2rem; font-weight:700; margin:8px 0 6px; color:#0f172a;'>LinkedIn Job Finder</h1>
        <p style='color:#64748b; font-size:0.95rem;'>Search smarter. Track everything. Apply with confidence.</p>
    </div>
    """, unsafe_allow_html=True)

    _, col, _ = st.columns([1, 1.4, 1])
    with col:
        mode = st.radio("Account", ["Sign In", "Create Account"], horizontal=True, label_visibility="collapsed")
        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

        if mode == "Create Account":
            username = st.text_input("Your Name", placeholder="e.g. Arjun Nair")
            email    = st.text_input("Email address", placeholder="arjun@gmail.com")
            pw       = st.text_input("Password", type="password", placeholder="Min 6 characters")
            pw2      = st.text_input("Confirm Password", type="password")
            if st.button("Create Account", type="primary", use_container_width=True):
                errors = []
                if not username.strip(): errors.append("Please enter your name.")
                if not email.strip():    errors.append("Please enter your email.")
                if not pw:               errors.append("Please enter a password.")
                if pw != pw2:            errors.append("Passwords do not match.")
                if pw and len(pw) < 6:   errors.append("Password must be at least 6 characters.")
                for e in errors: st.error(e)
                if not errors:
                    user, err = do_signup(email.strip(), pw, username.strip())
                    if err:   st.error(f"Sign up failed: {err}")
                    else:     st.success("✅ Account created! Check your email for a verification link, then Sign In.")
        else:
            email = st.text_input("Email address")
            pw    = st.text_input("Password", type="password")
            if st.button("Sign In", type="primary", use_container_width=True):
                if not email or not pw:
                    st.error("Please enter your email and password.")
                else:
                    res, err = do_signin(email.strip(), pw)
                    if err:
                        st.error(err)
                    else:
                        st.session_state.user          = res.user
                        st.session_state.access_token  = res.session.access_token
                        st.session_state.refresh_token = res.session.refresh_token
                        # Load saved filter preferences immediately after login
                        prefs = load_user_prefs(res.user.id)
                        for k, v in prefs.items():
                            if k not in st.session_state:
                                st.session_state[k] = v
                        st.session_state.prefs_loaded = True
                        st.rerun()

        st.markdown("""
        <div style='text-align:center; margin-top:20px; font-size:12px; color:#94a3b8; line-height:1.8;'>
            🔒 Your job searches are private and only visible to you.<br>
            The admin can see registered email addresses for access management.
        </div>
        """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────────────────────────
@st.dialog("⚠️ Required Fields Missing")
def show_missing_fields_dialog(missing_fields: list):
    st.markdown("""
    <div style='text-align:center; padding:8px 0 12px;'>
        <div style='font-size:2.5rem; margin-bottom:8px;'>📝</div>
        <div style='font-size:1.1rem; font-weight:600; color:#0f172a; margin-bottom:6px;'>
            Please fill in the required fields
        </div>
    </div>
    """, unsafe_allow_html=True)
    for field in missing_fields:
        st.error(f"**{field}** is required before searching.")
    st.markdown("""
    <div style='font-size:0.85rem; color:#64748b; margin-top:10px; line-height:1.7;'>
    These fields are mandatory to prevent overly broad searches that would return
    too many irrelevant results.<br><br>
    💡 <strong>Job Title</strong> — type the exact role you want, e.g. <em>Talent Acquisition Manager</em><br>
    💡 <strong>Locations</strong> — type at least one city or country, e.g. <em>India, Dubai</em>
    </div>
    """, unsafe_allow_html=True)
    if st.button("✅ Got it — let me fill them in", type="primary", use_container_width=True):
        st.rerun()


@st.dialog("⚠️ Search Too Wide")
def show_too_wide_dialog(new_count, combo_count):
    st.markdown(f"""
    <div style="text-align:center; padding: 8px 0 12px;">
        <div style="font-size:2.5rem; margin-bottom:8px;">🔍</div>
        <div style="font-size:1.1rem; font-weight:600; color:#0f172a; margin-bottom:6px;">
            Too many new results ({new_count} jobs)
        </div>
        <div style="font-size:0.88rem; color:#64748b; line-height:1.7;">
            Your search returned <strong>{new_count} new jobs</strong> after removing duplicates
            — the limit is <strong>100</strong>.<br>
            No results have been saved to your repository.
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.error(
        f"**{new_count} new jobs found across {combo_count} search combination(s)** — "
        "this exceeds the 100-job limit. Nothing has been saved."
    )

    st.markdown("**To fix this, make your search more specific:**")
    st.markdown("""
- 📍 **Reduce locations** — use 1–2 specific cities instead of a whole country
- 🏷️ **Be more specific with the job title** — e.g. *Talent Acquisition Manager* instead of *Talent Acquisition*
- 📅 **Shorten the date range** — try *Last 3 days* or *Last 7 days*
- 🎯 **Add Primary Keywords** — force the job to mention specific terms like *SaaS* or *analytics*
- 💼 **Reduce work modes** — e.g. select only *Remote* instead of Remote + Hybrid + Onsite
    """)

    st.info("Close this popup and adjust your filters, then search again. "
            "These jobs will be fetched fresh on your next search.")

    if st.button("✅ Got it — close and adjust filters", type="primary", use_container_width=True):
        st.rerun()


def show_main_app():
    user     = st.session_state.user
    username = (user.user_metadata or {}).get("username", user.email.split("@")[0])

    # ── Sidebar ──────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("## 🔍 Job Search Filters")

        # User badge + sign out
        st.markdown(f"""
        <div style='background:#1e293b; border-radius:8px; padding:10px 14px; margin-bottom:10px;'>
            <div style='font-size:13px; font-weight:500;'>👤 {username}</div>
            <div style='font-size:11px; color:#94a3b8;'>{user.email}</div>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Sign Out", use_container_width=True):
            do_signout()

        st.markdown("---")

        # ── Job Title ─────────────────────────────────────────────
        filter_label("🏷️", "Job Title",
            "The role you are searching for. You can add multiple titles separated by commas. "
            "e.g. Talent Acquisition Manager, HR Business Partner, TA Lead")
        job_title = st.text_input(
            "job_title_input", label_visibility="collapsed",
            value="Talent Acquisition Manager, Head of Recruiting",
            placeholder="e.g. Talent Acquisition Manager, TA Lead",
        )

        # ── Locations ─────────────────────────────────────────────
        filter_label("📍", "Locations (comma-separated)",
            "Type any city, country, or Remote — separated by commas. "
            "e.g. India, Dubai, Singapore, Remote, United States")
        locations_raw = st.text_input(
            "locations_input", label_visibility="collapsed",
            value="India, Kochi",
            placeholder="e.g. India, Dubai, Singapore, Remote",
        )
        locations = [l.strip() for l in locations_raw.split(",") if l.strip()]

        # ── Work Mode ─────────────────────────────────────────────
        filter_label("💼", "Work Mode",
            "Remote = fully from home · Hybrid = mix of office and home · Onsite = office only. "
            "Select one or more.")
        work_modes = st.multiselect(
            "work_mode_input", label_visibility="collapsed",
            options=["remote", "hybrid", "onsite"],
            default=["remote", "hybrid"],
        )

        # ── Experience Level ──────────────────────────────────────
        filter_label("📊", "Experience Level",
            "Seniority of roles to include. Mid-Senior suits 8+ years of experience. "
            "Select multiple to widen the search.")
        exp_levels = st.multiselect(
            "exp_level_input", label_visibility="collapsed",
            options=["internship", "entry", "associate", "mid-senior", "director", "executive", "any"],
            default=["mid-senior"],
        )

        # ── Job Type ──────────────────────────────────────────────
        filter_label("📋", "Job Type",
            "Full-time = permanent role · Contract = fixed term · Part-time · Any = all types.")
        job_type = st.selectbox(
            "job_type_input", label_visibility="collapsed",
            options=["full-time", "contract", "part-time", "any"],
            index=0,
        )

        # ── Posted Within ─────────────────────────────────────────
        filter_label("📅", "Posted Within",
            "How recently the job was posted on LinkedIn. Last 7 days gives fresh results. "
            "Use Last 30 days for more volume in niche roles.")
        posted_days = st.select_slider(
            "posted_days_input", label_visibility="collapsed",
            options=[1, 3, 7, 14, 30], value=7,
            format_func=lambda x: f"Last {x} day{'s' if x > 1 else ''}",
        )

        # ── Min Exp (full width, max_results removed from UI) ────
        filter_label("👤", "Min Exp (yrs)",
            "Jobs requiring fewer years than this will be filtered out. Set to 0 to see all.")
        min_exp = st.number_input(
            "min_exp_input", label_visibility="collapsed",
            min_value=0, max_value=30, value=0, step=1,
        )
        max_results = 200  # Internal fetch cap — 100 new-job limit enforced after dedup

        st.markdown("---")

        # ── Primary Must-have Keywords ────────────────────────────
        with st.expander("🎯 Primary Keywords (AND — ALL must match)", expanded=False):
            filter_label("✅", "Primary Must-have",
                "ALL words listed here must appear in the job description. "
                "e.g. recruitment, hiring means the job must mention BOTH words. "
                "Leave empty to disable this filter.")
            primary_text = st.text_area(
                "primary_kw_input", label_visibility="collapsed",
                value="recruitment\nhiring",
                height=90,
                placeholder="One keyword per line",
            )

        # ── Secondary Must-have Keywords ──────────────────────────
        with st.expander("🎯 Secondary Keywords (OR — ANY one must match)", expanded=False):
            filter_label("🔍", "Secondary Must-have",
                "At least ONE word from this list must appear in the job description. "
                "e.g. SaaS, analytics, IT services means ANY one of these must be present. "
                "Leave empty to disable.")
            secondary_text = st.text_area(
                "secondary_kw_input", label_visibility="collapsed",
                value="analytics",
                height=90,
                placeholder="One keyword per line",
            )

        # ── Exclude Keywords ─────────────────────────────────────
        with st.expander("🚫 Exclude Keywords", expanded=False):
            filter_label("🚫", "Exclude Keywords",
                "Jobs mentioning ANY of these words are removed entirely from results. One per line. "
                "e.g. BPO, staffing, intern")
            exclude_text = st.text_area(
                "exclude_kw_input", label_visibility="collapsed",
                value="BPO\nstaffing\nRPO\njunior\nassociate recruiter\nfresher\nentry level\nintern\nrelocation required",
                height=110,
                placeholder="One keyword per line",
            )

        # ── Company Filter ────────────────────────────────────────
        with st.expander("🏢 Company Filter", expanded=False):
            filter_label("🏢", "Target Companies",
                "Leave blank to search all companies. Add company names (one per line) "
                "to restrict results to only those companies. Partial names work — "
                "'Fractal' will match 'Fractal Analytics', 'Fractal AI', etc.")
            companies_text = st.text_area(
                "companies_input", label_visibility="collapsed",
                value="", height=70,
                placeholder="e.g.\nFractal\nInfosys\nTata",
            )
            filter_label("🔒", "Hide Closed Jobs",
                "Toggle ON to exclude jobs that are no longer accepting applications.")
            exclude_closed = st.toggle(
                "exclude_closed_input", label_visibility="collapsed", value=True,
            )

        # ── Email ─────────────────────────────────────────────────
        with st.expander("📧 Email Results (optional)", expanded=False):
            st.caption("Uses Gmail. Generate an App Password at myaccount.google.com → Security.")
            email_from     = st.text_input("From (Gmail)", "")
            email_password = st.text_input("App Password", "", type="password")
            email_to       = st.text_input("Send To", "")

        st.markdown("---")
        run = st.button("🚀 Search Jobs", type="primary", use_container_width=True)
        st.markdown("""
        <div style='color:#475569; font-size:0.72rem; margin-top:12px; line-height:1.7;'>
        ℹ️ Results are deduplicated against your repository — jobs already found in previous searches are skipped automatically.
        </div>
        """, unsafe_allow_html=True)

    # ── Tabs ─────────────────────────────────────────────────────
    tab_search, tab_repo = st.tabs(["🔍 Search Results", "📁 My Job Repository"])

    # ═══════════════════════════════════════════════════════════
    # TAB 1 — SEARCH
    # ═══════════════════════════════════════════════════════════
    with tab_search:
        st.markdown(f"""
        <div style='display:flex; align-items:center; gap:12px; margin-bottom:4px;'>
          <span style='font-size:2rem;'>🔍</span>
          <h1 style='font-size:1.8rem; font-weight:700; color:#0f172a; margin:0;'>LinkedIn Job Finder</h1>
        </div>
        <p style='color:#64748b; margin-bottom:20px;'>Welcome, <strong>{username}</strong>! Configure filters in the sidebar and click Search Jobs.</p>
        """, unsafe_allow_html=True)

        if run:
            missing = []
            if not job_title.strip():
                missing.append("Job Title")
            if not locations:
                missing.append("Locations")
            if missing:
                show_missing_fields_dialog(missing)
                st.stop()
            errors = []
            if not work_modes: errors.append("Select at least one Work Mode.")
            if not exp_levels: errors.append("Select at least one Experience Level.")
            for e in errors: st.error(e)
            if errors: st.stop()

            # Parse job titles — support comma-separated
            job_titles = [t.strip() for t in job_title.split(",") if t.strip()]
            primary_kws   = [k.strip() for k in primary_text.splitlines()   if k.strip()]
            secondary_kws = [k.strip() for k in secondary_text.splitlines() if k.strip()]
            exclude_kws   = [k.strip() for k in exclude_text.splitlines()   if k.strip()]
            target_cos    = [c.strip() for c in companies_text.splitlines() if c.strip()]

            all_jobs = []

            log_lines  = []
            log_holder = st.empty()
            def ui_log(msg):
                log_lines.append(msg)
                rendered = "\n".join(log_lines[-18:])
                log_holder.markdown(f"<div class='log-box'>{rendered}</div>", unsafe_allow_html=True)
            lj.log = ui_log

            with st.status("Searching LinkedIn…", expanded=True) as status:
                prog = st.progress(0, text="Starting…")

                # Build shared cfg (same for all titles)
                cfg_base = {
                    "location":             locations,
                    "work_mode":            work_modes,
                    "experience_level":     exp_levels,
                    "job_type":             job_type,
                    "posted_within_days":   posted_days,
                    "min_years_experience": int(min_exp),
                    "max_results":          max_results,
                    "primary_must_have":    primary_kws,
                    "secondary_must_have":  secondary_kws,
                    "preferred_keywords":   [],
                    "exclude_keywords":     exclude_kws,
                    "target_companies":     target_cos,
                    "exclude_closed_jobs":  exclude_closed,
                    "skills":               [],
                    "output_folder":        ".",
                    "output_filename":      "linkedin_jobs_web.xlsx",
                    "email_from":           email_from if "email_from" in dir() else "",
                    "email_password":       email_password if "email_password" in dir() else "",
                    "email_to":             email_to if "email_to" in dir() else "",
                    "email_smtp":           "smtp.gmail.com",
                    "email_port":           587,
                    "schedule_time":        "10:00",
                }

                # ── Step 1: Fetch all cards across all titles ────────────
                raw_fetched = []
                for ti, jt in enumerate(job_titles):
                    cfg = {**cfg_base, "job_title": jt}
                    st.write(f"**Step 1/{len(job_titles)+2} — Fetching:** {jt}")
                    fetched = lj.fetch_all_combinations(cfg)
                    raw_fetched.extend(fetched)
                    prog.progress(int(25 * (ti + 1) / len(job_titles)), text=f"{len(raw_fetched)} raw listings fetched…")

                # ── Step 2: Dedup across titles AND against repository ────
                st.write(f"**Step {len(job_titles)+1}/{len(job_titles)+2} — Deduplicating {len(raw_fetched)} listings…**")
                seen_urls_local = set()
                unique_fetched = []
                for j in raw_fetched:
                    u = j.get("LinkedIn URL", "")
                    if u and u not in seen_urls_local:
                        seen_urls_local.add(u)
                        unique_fetched.append(j)
                    elif not u:
                        unique_fetched.append(j)

                # Check against Supabase repository
                existing_repo_urls = get_existing_urls(user.id)
                new_only = [j for j in unique_fetched if j.get("LinkedIn URL", "") not in existing_repo_urls]
                already_seen = len(unique_fetched) - len(new_only)

                prog.progress(35, text=f"{len(new_only)} new · {already_seen} already in repository — skipping enrichment on duplicates…")
                st.write(f"  ✅ {len(new_only)} new jobs to enrich · ⏭️ {already_seen} already in your repository — skipped")

                if not new_only:
                    status.update(label="No new jobs found — all results already in your repository", state="complete")
                    st.session_state.results   = []
                    st.session_state.new_count = 0
                    st.session_state.dup_count = already_seen
                    st.stop()

                # ── Cap check: >100 new jobs → show popup, abort ─────────
                RESULT_CAP = 100
                if len(new_only) > RESULT_CAP:
                    status.update(label=f"⚠️ Search too wide — {len(new_only)} new jobs found (limit: {RESULT_CAP})", state="error")
                    combo_count = len(locations) * len(work_modes) * len(exp_levels) * len(job_titles)
                    # Clear any previous results so stale data is not shown
                    st.session_state.results   = None
                    st.session_state.new_count = 0
                    st.session_state.dup_count = 0
                    show_too_wide_dialog(len(new_only), combo_count)
                    st.stop()

                # ── Step 3: Enrich only the new jobs ────────────────────
                st.write(f"**Step {len(job_titles)+2}/{len(job_titles)+2} — Enriching {len(new_only)} new jobs** (Easy Apply · score · company info)…")
                st.caption("~2 seconds per job — only new jobs are enriched, duplicates skipped.")
                cfg_enrich = {**cfg_base, "job_title": job_titles[0]}
                all_jobs = lj.enrich_jobs(new_only, cfg_enrich)
                prog.progress(90, text="Saving to repository…")

                if not all_jobs:
                    status.update(label="No results passed filters", state="error")
                    st.warning("No jobs passed the filters after enrichment. Try adjusting your keyword or experience filters.")
                    st.stop()

                st.write("Saving to your repository…")
                new_count, dup_count = append_new_jobs(all_jobs, user)
                total_dup = already_seen + dup_count

                # Persist the filters used in this search for next login
                save_user_prefs(user.id, {
                    "job_title_input":      job_title,
                    "locations_input":      locations_raw,
                    "work_mode_input":      work_modes,
                    "exp_level_input":      exp_levels,
                    "job_type_input":       job_type,
                    "posted_days_input":    posted_days,
                    "min_exp_input":        int(min_exp),
                    "primary_kw_input":     primary_text,
                    "secondary_kw_input":   secondary_text,
                    "exclude_kw_input":     exclude_text,
                    "companies_input":      companies_text,
                })
                st.session_state.results   = all_jobs
                st.session_state.new_count = new_count
                st.session_state.dup_count = total_dup

                prog.progress(100, text="Done!")
                label = f"✅ Done! {new_count} new job(s) saved to your repository"
                if dup_count: label += f" · {dup_count} duplicate(s) skipped"
                status.update(label=label, state="complete")

            log_holder.empty()

        # ── Results panel ─────────────────────────────────────────
        if st.session_state.results:
            jobs = st.session_state.results
            df   = pd.DataFrame(jobs)
            for col in lj.COL_ORDER:
                if col not in df.columns: df[col] = ""
            df = df[lj.COL_ORDER]
            # Normalise Company Size so column sorts correctly (A→H buckets)
            if "Company Size" in df.columns:
                df["Company Size"] = df["Company Size"].apply(sortable_company_size)

            new_c = st.session_state.new_count
            dup_c = st.session_state.dup_count

            if dup_c and dup_c > 0:
                st.info(f"**{new_c} new job(s)** added to your repository · **{dup_c} duplicate(s)** already existed — skipped.")
            elif new_c is not None:
                st.success(f"**{new_c} new job(s)** added to your repository. No duplicates found.")

            m1, m2, m3, m4, m5, m6 = st.columns(6)
            m1.metric("Total Found",    len(df))
            m2.metric("New to Repo",    new_c or 0)
            m3.metric("Duplicates",     dup_c or 0)
            score_col = "Score" if "Score" in df.columns else ("Match" if "Match" in df.columns else None)
            m4.metric("Score ≥ 80", len(df[pd.to_numeric(df[score_col], errors="coerce") >= 80]) if score_col else 0)
            m5.metric("Easy Apply",     len(df[df["Easy Apply"] == "Yes"]))
            m6.metric("Still Open",     len(df[df["Still Accepting?"] == "Yes"]))

            st.markdown("---")

            # Generation counter for clear-filter reset
            if "sr_filter_gen" not in st.session_state:
                st.session_state.sr_filter_gen = 0
            sr_gen = st.session_state.sr_filter_gen

            all_easy    = sorted(df["Easy Apply"].dropna().unique().tolist())
            all_open    = sorted(df["Still Accepting?"].dropna().unique().tolist())
            all_locs    = sorted(df["Location"].dropna().unique().tolist())
            all_p_match = sorted(df["Primary Keywords Match"].dropna().unique().tolist()) if "Primary Keywords Match" in df.columns else []

            filter_row_sr, btn_col_sr = st.columns([5, 1])
            with btn_col_sr:
                st.markdown("<div style='margin-top:4px'></div>", unsafe_allow_html=True)
                if st.button("🔄 Clear Filters", use_container_width=True,
                             help="Reset all filters to show all results", key="sr_clear"):
                    st.session_state.sr_filter_gen += 1
                    st.rerun()

            with filter_row_sr:
                f1, f2, f3, f4 = st.columns(4)
                sel_easy   = f1.multiselect("Easy Apply",       all_easy,    default=all_easy,    key=f"sr_e_{sr_gen}")
                sel_open   = f2.multiselect("Still Accepting",  all_open,    default=all_open,    key=f"sr_o_{sr_gen}")
                sel_loc    = f3.multiselect("Location",         all_locs,    default=all_locs,    key=f"sr_l_{sr_gen}")
                sel_pmatch = f4.multiselect("Primary KW Match", all_p_match, default=all_p_match, key=f"sr_pm_{sr_gen}")

            # Empty filter = show all for that dimension
            active_easy   = sel_easy   if sel_easy   else all_easy
            active_open   = sel_open   if sel_open   else all_open
            active_loc    = sel_loc    if sel_loc    else all_locs
            active_pmatch = sel_pmatch if sel_pmatch else all_p_match

            mask = (
                df["Easy Apply"].isin(active_easy) &
                df["Still Accepting?"].isin(active_open) &
                df["Location"].isin(active_loc)
            )
            if all_p_match:
                mask = mask & df["Primary Keywords Match"].isin(active_pmatch)
            fdf = df[mask].copy()

            filter_active_sr = (sel_easy != all_easy or sel_open != all_open
                                or sel_loc != all_locs or sel_pmatch != all_p_match)
            sr_filter_note = " · 🔽 Filters active" if filter_active_sr else ""
            st.caption(f"Showing **{len(fdf)}** of **{len(df)}** results{sr_filter_note}")

            st.dataframe(
                fdf, use_container_width=True, hide_index=True, height=460,
                column_config={
                    "Applied?":                   st.column_config.SelectboxColumn("Applied?",
                        options=["", "Relevant", "Applied", "Irrelevant - Criteria mismatch"], width="medium"),
                    "Job Title":                  st.column_config.TextColumn("Job Title",   width="large"),
                    "Company":                    st.column_config.TextColumn("Company",     width="medium"),
                    "Company Size":               st.column_config.TextColumn("Co. Size",    width="small"),
                    "Company Type":               st.column_config.TextColumn("Co. Type",    width="medium"),
                    "Level":                      st.column_config.TextColumn("Level",       width="small"),
                    "Location":                   st.column_config.TextColumn("Location",    width="small"),
                    "Posted Date":                st.column_config.TextColumn("Posted",      width="small"),
                    "Easy Apply":                 st.column_config.TextColumn("Easy Apply",  width="small"),
                    "Recruiter Profile":          st.column_config.LinkColumn("Recruiter",   display_text="👤", width="small"),
                    **({"Score": st.column_config.NumberColumn("Score", width="small", format="%d")} if "Score" in fdf.columns else {}),
                    **({"Match": st.column_config.TextColumn("Match", width="small")} if "Match" in fdf.columns else {}),
                    "Primary Keywords Match":     st.column_config.TextColumn("Primary KW",  width="small"),
                    "Secondary Keywords Match":   st.column_config.TextColumn("Secondary KW",width="small"),
                    "Work Mode":                  st.column_config.TextColumn("Mode",        width="small"),
                    "Still Accepting?":           st.column_config.TextColumn("Open?",       width="small"),
                    "Date Searched":              st.column_config.TextColumn("Searched",    width="small"),
                    "Exp. Required":              st.column_config.TextColumn("Exp. Req.",   width="small"),
                    "LinkedIn URL":               st.column_config.LinkColumn("LinkedIn",    display_text="🔗 View", width="small"),
                },
            )

            st.markdown("---")
            dl1, dl2, _ = st.columns([1, 1, 2])
            ts = datetime.now().strftime("%Y%m%d_%H%M")
            dl1.download_button("⬇️ Download CSV", fdf.to_csv(index=False).encode("utf-8"),
                f"jobs_{ts}.csv", "text/csv", use_container_width=True)
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as w:
                fdf.to_excel(w, index=False, sheet_name="LinkedIn Jobs")
                lj.apply_styles_and_dropdown(w.sheets["LinkedIn Jobs"], len(fdf))
            buf.seek(0)
            dl2.download_button("⬇️ Download Excel", buf, f"jobs_{ts}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
            st.caption("💡 Excel is colour-coded: green = Score ≥ 80, blue = Score 50–79. Applied? dropdown included.")

        elif not run:
            st.markdown("""
            <div style='text-align:center; padding:60px 0; color:#94a3b8;'>
              <div style='font-size:4rem; margin-bottom:16px;'>📋</div>
              <div style='font-size:1.2rem; font-weight:600; color:#475569; margin-bottom:8px;'>
                Configure filters in the sidebar and click Search Jobs
              </div>
              <div style='font-size:0.9rem;'>
                Hover over the <b>ℹ</b> icon next to each filter for guidance.<br>
                New jobs are saved to your repository automatically. Duplicates are skipped.
              </div>
            </div>
            """, unsafe_allow_html=True)

    # ═══════════════════════════════════════════════════════════
    # TAB 2 — REPOSITORY
    # ═══════════════════════════════════════════════════════════
    with tab_repo:
        st.markdown("""
        <div style='display:flex; align-items:center; gap:12px; margin-bottom:4px;'>
          <span style='font-size:2rem;'>📁</span>
          <h1 style='font-size:1.8rem; font-weight:700; color:#0f172a; margin:0;'>My Job Repository</h1>
        </div>
        <p style='color:#64748b; margin-bottom:20px;'>
          Every job found across all your searches — deduplicated and persisted.
          Update the <strong>Status ✏️</strong> column by clicking any cell.
        </p>
        """, unsafe_allow_html=True)

        with st.spinner("Loading your repository…"):
            repo_data = get_repository(user.id)

        if not repo_data:
            st.markdown("""
            <div style='text-align:center; padding:60px 0; color:#94a3b8;'>
              <div style='font-size:4rem; margin-bottom:16px;'>📭</div>
              <div style='font-size:1.2rem; font-weight:600; color:#475569; margin-bottom:8px;'>Your repository is empty</div>
              <div style='font-size:0.9rem;'>Run a search and your jobs will appear here automatically.</div>
            </div>
            """, unsafe_allow_html=True)
            return

        repo_df = pd.DataFrame(repo_data)
        # Normalise Company Size for correct sort order in repository table
        if "company_size" in repo_df.columns:
            repo_df["company_size"] = repo_df["company_size"].apply(sortable_company_size)

        # Metrics
        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("Total Jobs", len(repo_df))
        m2.metric("New",        len(repo_df[repo_df["status"] == "New"]))
        m3.metric("Relevant",   len(repo_df[repo_df["status"] == "Relevant"]))
        m4.metric("Applied",    len(repo_df[repo_df["status"] == "Applied"]))
        m5.metric("Irrelevant", len(repo_df[repo_df["status"] == "Irrelevant"]))

        st.markdown("---")

        # Reset-based filters
        if "repo_filter_gen" not in st.session_state:
            st.session_state.repo_filter_gen = 0
        gen = st.session_state.repo_filter_gen

        all_statuses = ["New", "Relevant", "Applied", "Irrelevant"]
        all_r_locs   = sorted(repo_df["location"].dropna().unique().tolist())
        all_r_comps  = sorted(repo_df["company"].dropna().unique().tolist())

        # Score bucket filter
        def score_bucket(v):
            try:
                s = int(float(str(v)))
                if s >= 80: return "High (≥80)"
                if s >= 50: return "Medium (50–79)"
                return "Low (<50)"
            except: return "NA"
        if "score" in repo_df.columns:
            repo_df["_score_bucket"] = repo_df["score"].apply(score_bucket)
        else:
            repo_df["_score_bucket"] = "NA"
        all_buckets = sorted(repo_df["_score_bucket"].unique().tolist())

        filter_row, btn_col = st.columns([5, 1])
        with btn_col:
            st.markdown("<div style='margin-top:4px'></div>", unsafe_allow_html=True)
            if st.button("🔄 Clear Filters", use_container_width=True, help="Reset all filters to show full repository"):
                st.session_state.repo_filter_gen += 1
                st.rerun()

        with filter_row:
            rf1, rf2, rf3, rf4 = st.columns(4)
            sel_st  = rf1.multiselect("Status",   all_statuses, default=all_statuses,  key=f"rp_st_{gen}")
            sel_rl  = rf2.multiselect("Location", all_r_locs,   default=all_r_locs,    key=f"rp_l_{gen}")
            sel_rc  = rf3.multiselect("Company",  all_r_comps,  default=all_r_comps,   key=f"rp_c_{gen}")
            sel_sb  = rf4.multiselect("Score",    all_buckets,  default=all_buckets,   key=f"rp_sb_{gen}")

        active_st = sel_st if sel_st else all_statuses
        active_rl = sel_rl if sel_rl else all_r_locs
        active_rc = sel_rc if sel_rc else all_r_comps
        active_sb = sel_sb if sel_sb else all_buckets

        rmask = (
            repo_df["status"].isin(active_st) &
            repo_df["location"].isin(active_rl) &
            repo_df["company"].isin(active_rc) &
            repo_df["_score_bucket"].isin(active_sb)
        )
        filtered_repo = repo_df[rmask].copy().reset_index(drop=True)
        filter_active = (sel_st != all_statuses or sel_rl != all_r_locs
                         or sel_rc != all_r_comps or sel_sb != all_buckets)
        filter_note = " · 🔽 Filters active" if filter_active else ""
        st.caption(f"Showing **{len(filtered_repo)}** of **{len(repo_df)}** jobs  ·  Click any **Status ✏️** cell to update{filter_note}")

        DISPLAY_COLS = [
            "status", "job_title", "company", "company_size", "company_type",
            "level", "location", "work_mode", "score",
            "primary_keywords_match", "secondary_keywords_match",
            "easy_apply", "still_accepting", "posted_date",
            "exp_required", "recruiter_profile", "linkedin_url", "date_added",
        ]
        for col in DISPLAY_COLS:
            if col not in filtered_repo.columns: filtered_repo[col] = ""
        display_df = filtered_repo[DISPLAY_COLS].copy()

        edited_df = st.data_editor(
            display_df,
            use_container_width=True,
            hide_index=True,
            height=520,
            key="repo_editor",
            column_config={
                "status":                    st.column_config.SelectboxColumn("Status ✏️",
                    options=["New", "Relevant", "Applied", "Irrelevant"], width="medium", required=True),
                "job_title":                 st.column_config.TextColumn("Job Title",        width="large"),
                "company":                   st.column_config.TextColumn("Company",          width="medium"),
                "company_size":              st.column_config.TextColumn("Co. Size",         width="small"),
                "company_type":              st.column_config.TextColumn("Co. Type",         width="medium"),
                "level":                     st.column_config.TextColumn("Level",            width="small"),
                "location":                  st.column_config.TextColumn("Location",         width="small"),
                "work_mode":                 st.column_config.TextColumn("Mode",             width="small"),
                "score":                     st.column_config.TextColumn("Score",            width="small"),
                "primary_keywords_match":    st.column_config.TextColumn("Primary KW",       width="small"),
                "secondary_keywords_match":  st.column_config.TextColumn("Secondary KW",     width="small"),
                "easy_apply":                st.column_config.TextColumn("Easy Apply",       width="small"),
                "still_accepting":           st.column_config.TextColumn("Open?",            width="small"),
                "posted_date":               st.column_config.TextColumn("Posted",           width="small"),
                "exp_required":              st.column_config.TextColumn("Exp. Req.",        width="small"),
                "recruiter_profile":         st.column_config.LinkColumn("Recruiter",        display_text="👤", width="small"),
                "linkedin_url":              st.column_config.LinkColumn("LinkedIn",         display_text="🔗 View", width="small"),
                "date_added":                st.column_config.TextColumn("Added On",         width="medium"),
            },
            disabled=[c for c in DISPLAY_COLS if c != "status"],
        )

        # Persist status changes
        orig_status   = display_df["status"].reset_index(drop=True)
        edited_status = edited_df["status"].reset_index(drop=True)
        changed_mask  = orig_status != edited_status
        if changed_mask.any():
            for idx in changed_mask[changed_mask].index.tolist():
                update_job_status(filtered_repo.iloc[idx]["id"], edited_df.iloc[idx]["status"])
            st.success(f"✅ {changed_mask.sum()} status update(s) saved.")
            st.rerun()

        st.markdown("---")

        # Downloads
        dl1, dl2, _ = st.columns([1, 1, 2])
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        RENAME = {
            "status": "Status", "job_title": "Job Title", "company": "Company",
            "company_size": "Company Size", "company_type": "Company Type",
            "level": "Level", "location": "Location", "work_mode": "Work Mode",
            "score": "Score", "primary_keywords_match": "Primary Keywords Match",
            "secondary_keywords_match": "Secondary Keywords Match",
            "easy_apply": "Easy Apply", "still_accepting": "Still Accepting?",
            "posted_date": "Posted Date", "exp_required": "Exp. Required",
            "recruiter_profile": "Recruiter Profile", "linkedin_url": "LinkedIn URL",
            "date_added": "Date Added",
        }
        download_df = filtered_repo[DISPLAY_COLS].copy().rename(columns=RENAME)

        dl1.download_button("⬇️ Download CSV",
            download_df.to_csv(index=False).encode("utf-8"),
            f"repository_{ts}.csv", "text/csv", use_container_width=True)

        buf2 = io.BytesIO()
        with pd.ExcelWriter(buf2, engine="openpyxl") as w:
            download_df.to_excel(w, index=False, sheet_name="Job Repository")
            ws = w.sheets["Job Repository"]
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.worksheet.datavalidation import DataValidation
            hdr = PatternFill("solid", fgColor="1F3864")
            for cell in ws[1]:
                cell.font      = Font(bold=True, color="FFFFFF", size=10)
                cell.fill      = hdr
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.freeze_panes = "A2"
            for col_cells in ws.columns:
                w_ = max((len(str(c.value or "")) for c in col_cells), default=10)
                ws.column_dimensions[col_cells[0].column_letter].width = min(w_ + 4, 45)
            dv = DataValidation(type="list", formula1='"New,Relevant,Applied,Irrelevant"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.sqref = f"A2:A{len(download_df) + 1}"
        buf2.seek(0)
        dl2.download_button("⬇️ Download Excel", buf2,
            f"repository_{ts}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
        st.caption("💡 Status column is editable directly in the table above — changes save instantly.")


# ─────────────────────────────────────────────────────────────────
# ROUTER
# ─────────────────────────────────────────────────────────────────
if st.session_state.user is None:
    show_auth_page()
else:
    show_main_app()
