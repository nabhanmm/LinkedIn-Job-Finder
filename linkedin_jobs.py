"""
LinkedIn Job Scheduler — v5 (final)
Easy Apply: detected via data-tracking-control-name="apply-link-onsite"
            confirmed from real LinkedIn page HTML analysis.
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import schedule
import time
import os, re, json, hashlib, smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# ═══════════════════════════════════════════════════════════════════
# CONFIGURATION — the only section you ever need to edit
# ═══════════════════════════════════════════════════════════════════
CONFIG = {
    # Job title — add multiple designations separated by commas inside double quotes
    # Example: "Talent Acquisition Manager", "Head of Recruiting", "TA Lead"
    # LinkedIn will search each as a broad keyword match.
    "job_title": ["Talent Acquisition Manager", "Head of Recruiting", "TA Lead", "Recruitment Manager", "HR Manager"],

    # ── Primary Must-have Keywords (AND logic) ─────────────────────
    # ALL words listed here must appear somewhere in the job title or description.
    # Use a comma-separated list. Leave as [] to disable this filter.
    # Example: ["analytics", "ATS"] means the job must mention BOTH words.
    "primary_must_have": ["recruitment", "hiring"],

    # ── Secondary Must-have Keywords (OR logic) ────────────────────
    # At least ONE word from this list must appear in the job title or description.
    # Works as an AND condition alongside primary_must_have.
    # Leave as [] to disable this filter.
    # Example: ["SaaS", "analytics", "IT services"] means the job must mention
    # at least one of these — but not necessarily all of them.
    "secondary_must_have": ["analytics"],
    "location": ["India", "Kochi", "Bangkok", "Thailand", "US"],
    "work_mode": ["remote"],
    "experience_level": ["any"],
    "job_type": "full-time",
    "posted_within_days": 30,
    "min_years_experience": 0,
    "preferred_keywords": [
        "analytics", "data analytics", "IT services", "SaaS",
        "consulting", "product", "technology", "AI", "machine learning"
    ],
    "exclude_keywords": [
        "BPO", "staffing", "RPO", "junior"
    ],

    # ── Company filter (optional) ──────────────────────────────────
    # List specific companies to search within. Partial names work —
    # "Fractal" will match "Fractal Analytics", "Fractal AI", etc.
    # Set to [] to search all companies (default behaviour).
    "target_companies": [],

    # ── Closed jobs filter ─────────────────────────────────────────
    # True  = exclude jobs that are no longer accepting applications (default)
    # False = include closed jobs (they will be flagged in the Excel)
    "exclude_closed_jobs": False,

    "skills": [
        "ATS", "sourcing", "recruitment", "talent acquisition",
        "hiring", "Excel", "Power Automate", "Power BI"
    ],
    "max_results":     100,
    "output_folder":   ".",
    "output_filename": "linkedin_ta_jobs_ALL.xlsx",
    "schedule_time":   "10:00",
    "email_from":      "",
    "email_password":  "",
    "email_to":        "",
    "email_smtp":      "smtp.gmail.com",
    "email_port":      587,
}
# ═══════════════════════════════════════════════════════════════════

WORK_MODE_MAP = {"remote": "2", "onsite": "1", "hybrid": "3", "any": ""}
JOB_TYPE_MAP  = {"full-time": "F", "contract": "C", "part-time": "P", "any": ""}
EXP_MAP = {
    "internship": "1", "entry": "2", "associate": "3",
    "mid-senior": "4", "director": "5", "executive": "6", "any": ""
}
TIME_MAP = {1: "r86400", 3: "r259200", 7: "r604800", 14: "r1209600", 30: "r2592000"}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

APPLIED_OPTIONS = ["", "Relevant", "Applied", "Irrelevant - Criteria mismatch"]

LEVEL_LABELS = {
    "internship": "Internship", "entry": "Entry Level", "associate": "Associate",
    "mid-senior": "Mid-Senior", "director": "Director", "executive": "Executive", "any": "Any",
}

COL_ORDER = [
    "Applied?",
    "Job Title", "Company", "Company Size", "Company Type",
    "Level", "Location", "Posted Date", "Easy Apply", "Recruiter Profile",
    "Score", "Primary Keywords Match", "Secondary Keywords Match",
    "Work Mode", "Still Accepting?", "Date Searched", "Exp. Required", "LinkedIn URL"
]


# ─────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────
def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)

def clean(el):
    return el.get_text(strip=True) if el else ""

def as_list(val):
    return val if isinstance(val, list) else [val]

def job_fingerprint(job):
    url = job.get("LinkedIn URL", "")
    if url:
        m = re.search(r'/(\d{10,})', url)
        if m:
            return m.group(1)
    raw = (job.get("Job Title", "") + job.get("Company", "")).lower().strip()
    return hashlib.md5(raw.encode()).hexdigest()[:16]

def safe_get(url, timeout=14):
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout)
        return r if r.status_code == 200 else None
    except Exception:
        return None


# ─────────────────────────────────────────────────────────────────
# Job detail page
# Easy Apply detection — confirmed from real LinkedIn HTML:
#   Easy Apply jobs have: data-tracking-control-name="public_jobs_apply-link-onsite"
#   External apply jobs:  data-tracking-control-name="public_jobs_apply-link-offsite"
#   The button text is just "Apply" in both cases — never "Easy Apply"
# ─────────────────────────────────────────────────────────────────
def fetch_job_detail(url):
    result = {"description": "", "easy_apply": False, "recruiter": ""}
    if not url:
        return result
    r = safe_get(url.split("?")[0])
    if not r:
        return result
    raw  = r.text
    soup = BeautifulSoup(raw, "html.parser")

    # Description
    desc_el = soup.find("div", class_=re.compile(r"description", re.I))
    result["description"] = desc_el.get_text(" ", strip=True) if desc_el else ""

    # Easy Apply — the only reliable signal on the public page
    result["easy_apply"] = "apply-link-onsite" in raw

    # Recruiter profile
    recruiter = ""
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if "/in/" in href and "linkedin.com" in href:
            m = re.search(r'(https?://[a-z.]*linkedin\.com/in/[^/?&#]+)', href)
            if m:
                recruiter = m.group(1)
                break
    if not recruiter:
        m = re.search(r'linkedin\.com/in/([a-zA-Z0-9\-]+)', raw)
        if m:
            recruiter = "https://www.linkedin.com/in/" + m.group(1)
    result["recruiter"] = recruiter

    # ── Closed / no longer accepting ─────────────────────────────
    # LinkedIn shows a banner: "No longer accepting applications"
    # It also appears as a class or aria text on the page.
    closed_signals = [
        "no longer accepting applications",
        "not accepting applications",
        "closed for applications",
        "job is no longer available",
        "this job is closed",
    ]
    raw_lower = raw.lower()
    result["is_closed"] = any(sig in raw_lower for sig in closed_signals)

    return result


# ─────────────────────────────────────────────────────────────────
# Company page scraper
# ─────────────────────────────────────────────────────────────────
def get_company_page_url(job_url):
    if not job_url:
        return None
    r = safe_get(job_url)
    if not r:
        return None
    soup = BeautifulSoup(r.text, "html.parser")
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if re.search(r'linkedin\.com/company/[^/?]+', href):
            clean_url = re.search(r'(https?://[a-z.]*linkedin\.com/company/[^/?]+)', href)
            if clean_url:
                return clean_url.group(1)
    m = re.search(r'(https?://[a-z.]*linkedin\.com/company/[^/"?]+)', r.text)
    return m.group(1) if m else None


def scrape_company_bio(company_url):
    result = {"company_size": "Not listed", "company_type": "Not listed"}
    if not company_url:
        return result
    for url in [company_url.rstrip("/") + "/about", company_url]:
        r = safe_get(url)
        if not r:
            continue
        text = r.text
        soup = BeautifulSoup(text, "html.parser")

        # Company size
        size_el = soup.find(attrs={"data-test-id": re.compile(r"about-us__size", re.I)})
        if not size_el:
            for dt in soup.find_all("dt"):
                if "company size" in dt.get_text(strip=True).lower():
                    dd = dt.find_next_sibling("dd")
                    if dd:
                        size_el = dd
                        break
        if not size_el:
            m = re.search(r'(\d[\d,]*\+?\s*(?:to\s*\d[\d,]*\+?)?\s*employees)', text, re.I)
            if m:
                result["company_size"] = m.group(1).strip()
        else:
            raw_size = size_el.get_text(strip=True)
            raw_size = re.sub(r'(?i)^company\s*size\s*:?\s*', '', raw_size).strip()
            result["company_size"] = raw_size

        # Company type / industry
        type_el = soup.find(attrs={"data-test-id": re.compile(r"about-us__industry|about-us__type", re.I)})
        if not type_el:
            for dt in soup.find_all("dt"):
                label = dt.get_text(strip=True).lower()
                if "industry" in label or "type" in label:
                    dd = dt.find_next_sibling("dd")
                    if dd:
                        type_el = dd
                        break
        if not type_el:
            for script in soup.find_all("script", type="application/ld+json"):
                try:
                    data = json.loads(script.string or "")
                    if isinstance(data, dict):
                        industry = data.get("industry") or data.get("knowsAbout")
                        if industry:
                            result["company_type"] = str(industry)
                            break
                except Exception:
                    pass
        else:
            raw_type = type_el.get_text(strip=True)
            raw_type = re.sub(r'(?i)^industry\s*:?\s*', '', raw_type).strip()
            result["company_type"] = raw_type

        if result["company_size"] != "Not listed" or result["company_type"] != "Not listed":
            break
        time.sleep(1)
    return result


# ─────────────────────────────────────────────────────────────────
# Experience filter
# ─────────────────────────────────────────────────────────────────
def extract_min_years(text):
    if not text:
        return None
    t = text.lower()
    found = []
    for m in re.finditer(
        r'(?:minimum|at\s+least|min\.?|over|more\s+than)?\s*(\d{1,2})\s*\+?\s*(?:or\s+more\s+)?years?', t
    ):
        found.append(int(m.group(1)))
    for m in re.finditer(r'(\d{1,2})\s*(?:-|to)\s*\d{1,2}\s*years?', t):
        found.append(int(m.group(1)))
    found = [y for y in found if 1 <= y <= 30]
    return min(found) if found else None


# ─────────────────────────────────────────────────────────────────
# Search — fetch cards from LinkedIn guest API
# ─────────────────────────────────────────────────────────────────
def build_params(location, work_mode, exp_level, cfg, start=0):
    tpr = TIME_MAP.get(cfg["posted_within_days"], TIME_MAP[7])
    params = {"keywords": cfg["job_title"], "location": location,
              "f_TPR": tpr, "start": start, "sortBy": "DD"}
    if wt  := WORK_MODE_MAP.get(work_mode, ""):      params["f_WT"] = wt
    if jt  := JOB_TYPE_MAP.get(cfg["job_type"], ""): params["f_JT"] = jt
    if exp := EXP_MAP.get(exp_level, ""):             params["f_E"]  = exp
    return params


def parse_card(card, exclude, preferred, exp_level="", work_mode_label=""):
    try:
        title   = clean(card.find("h3", class_="base-search-card__title"))
        company = clean(card.find("h4", class_="base-search-card__subtitle"))
        loc     = clean(card.find("span", class_="job-search-card__location"))
        time_el = card.find("time")
        link_el = card.find("a", class_="base-card__full-link")
        posted  = time_el.get("datetime", "")[:10] if time_el else ""
        url     = link_el.get("href", "").split("?")[0] if link_el else ""
        card_html = str(card)

        # ── Secondary remote validation ───────────────────────────
        # LinkedIn's guest API f_WT filter is "best effort" — it can
        # return on-site jobs when remote is requested. We do a second
        # check here using the card's own HTML metadata.
        # If the search asked for remote/hybrid but the card signals
        # on-site only, we drop it.
        if work_mode_label in ("Remote", "Hybrid"):
            # LinkedIn embeds workplace type in the card as a metadata span
            # or in data attributes. Check for explicit on-site signals.
            card_lower = card_html.lower()
            has_remote_signal = any(s in card_lower for s in [
                "remote", "hybrid", "work from home", "wfh", "telecommut"
            ])
            has_onsite_signal = any(s in card_lower for s in [
                "on-site", "onsite", "on site", "in-office", "in office"
            ])
            # Also check location string — if it's a city/country with no
            # "remote" mention in the location text itself, it may be on-site
            loc_lower = loc.lower()
            loc_is_remote = "remote" in loc_lower

            # Drop only if card explicitly signals on-site AND has no remote signal
            if has_onsite_signal and not has_remote_signal and not loc_is_remote:
                return None

        # Easy Apply not available in card HTML — set blank, filled by enrich pass
        easy    = ""

        level    = LEVEL_LABELS.get(exp_level, exp_level.title() if exp_level else "")
        combined = (title + " " + company + " " + loc).lower()

        for ex in exclude:
            if ex in combined:
                return None

        # NOTE: primary_must_have and secondary_must_have are applied
        # during the enrichment step where the full description is available.
        # Checking them here (card only has title+company+location) would
        # filter out almost everything incorrectly.

        return {
            "Applied?":                   "",
            "Job Title":                  title,
            "Company":                    company,
            "Company Size":               "",
            "Company Type":               "",
            "Level":                      level,
            "Location":                   loc,
            "Posted Date":                posted,
            "Easy Apply":                 easy,
            "Recruiter Profile":          "",
            "Score":                      "",   # filled by enrich pass
            "Primary Keywords Match":     "",   # filled by enrich pass
            "Secondary Keywords Match":   "",   # filled by enrich pass
            "Work Mode":                  work_mode_label,
            "Still Accepting?":           "",   # filled by enrich pass
            "Date Searched":              datetime.now().strftime("%Y-%m-%d %H:%M"),
            "Exp. Required":              "",
            "LinkedIn URL":               url,
        }
    except Exception:
        return None


def fetch_one_combo(location, work_mode, exp_level, cfg):
    base      = "https://www.linkedin.com/jobs-guest/jobs/api/seeMoreJobPostings/search"
    exclude   = [k.lower() for k in cfg["exclude_keywords"]]
    preferred = [k.lower() for k in cfg["preferred_keywords"]]
    wm_labels = {"remote": "Remote", "hybrid": "Hybrid", "onsite": "On-site", "any": ""}
    wm_label  = wm_labels.get(work_mode, "")
    jobs = []
    for start in range(0, cfg["max_results"], 25):
        try:
            r = requests.get(base,
                             params=build_params(location, work_mode, exp_level, cfg, start),
                             headers=HEADERS, timeout=15)
            if r.status_code != 200:
                break
            soup  = BeautifulSoup(r.text, "html.parser")
            cards = soup.find_all("div", class_="base-card")
            if not cards:
                break
            for card in cards:
                job = parse_card(card, exclude, preferred, exp_level, wm_label)
                if job:
                    jobs.append(job)
            if len(cards) < 25:
                break
        except Exception as e:
            log(f"    Request error: {e}")
            break
        time.sleep(1.5)
    return jobs


def company_matches(company_name, target_companies):
    """
    Fuzzy company match — checks if any target company name is a substring
    of the actual company name (case-insensitive).
    "Fractal" matches "Fractal Analytics", "Fractal AI Ltd", etc.
    """
    if not target_companies:
        return True   # no filter set — accept all companies
    company_lower = company_name.lower()
    return any(t.lower() in company_lower for t in target_companies)


def fetch_all_combinations(cfg):
    locations       = as_list(cfg["location"])
    work_modes      = as_list(cfg["work_mode"])
    exp_levels      = as_list(cfg["experience_level"])
    target_companies = cfg.get("target_companies", [])
    combos          = [(lo, wm, ex) for lo in locations for wm in work_modes for ex in exp_levels]

    primary_must_have   = [k.strip() for k in cfg.get("primary_must_have", []) if k.strip()]
    secondary_must_have = [k.strip() for k in cfg.get("secondary_must_have", []) if k.strip()]

    if target_companies:
        log(f"  Company filter active: {target_companies}")
    if primary_must_have:
        log(f"  Primary must-have (ALL required): {primary_must_have}")
    if secondary_must_have:
        log(f"  Secondary must-have (ANY ONE required): {secondary_must_have}")
    log(f"  Running {len(combos)} search combination(s)")

    seen_urls = {}
    for i, (lo, wm, ex) in enumerate(combos, 1):
        log(f"  [{i}/{len(combos)}] {lo} / {wm} / {ex}")
        for j in fetch_one_combo(lo, wm, ex, cfg):
            key = j.get("LinkedIn URL") or job_fingerprint(j)
            if key not in seen_urls:
                # Apply company filter — partial/fuzzy match
                if company_matches(j.get("Company", ""), target_companies):
                    seen_urls[key] = j
                else:
                    pass   # silently skip — company not in target list
        log(f"    → {len(seen_urls)} unique so far")
        time.sleep(2)
    return list(seen_urls.values())


# ─────────────────────────────────────────────────────────────────
# Deduplication (runs BEFORE enrich — saves time)
# ─────────────────────────────────────────────────────────────────
SEEN_FILE = "seen_jobs.json"

def load_seen():
    if os.path.exists(SEEN_FILE):
        try:
            with open(SEEN_FILE) as f:
                return set(json.load(f).get("ids", []))
        except Exception:
            pass
    return set()

def save_seen(ids):
    try:
        with open(SEEN_FILE, "w") as f:
            json.dump({"ids": list(ids),
                       "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M")}, f, indent=2)
    except Exception as e:
        log(f"  Warning: could not write seen_jobs.json — {e}")

def filter_new_only(jobs, cfg=None):
    seen = load_seen()
    excel_urls, excel_keys = set(), set()
    if cfg:
        fp = os.path.join(cfg.get("output_folder", "."), cfg.get("output_filename", ""))
        if os.path.exists(fp):
            try:
                ex = pd.read_excel(fp, sheet_name="LinkedIn Jobs",
                                   usecols=["Job Title", "Company", "LinkedIn URL"], dtype=str)
                ex.fillna("", inplace=True)
                excel_urls = set(ex["LinkedIn URL"].str.strip())
                excel_keys = set((r["Job Title"] + r["Company"]).lower().strip()
                                 for _, r in ex.iterrows())
            except Exception as e:
                log(f"  Note: could not read Excel for dedup ({e})")
    new_jobs, new_ids = [], set()
    for j in jobs:
        fid = job_fingerprint(j)
        url = str(j.get("LinkedIn URL", "")).strip()
        key = (str(j.get("Job Title","")) + str(j.get("Company",""))).lower().strip()
        if fid not in seen and url not in excel_urls and key not in excel_keys:
            new_jobs.append(j)
            new_ids.add(fid)
    seen.update(new_ids)
    save_seen(seen)
    skipped = len(jobs) - len(new_jobs)
    if skipped:
        log(f"  Deduplication: {skipped} already seen — skipped")
    log(f"  New jobs to enrich: {len(new_jobs)}")
    return new_jobs


# ─────────────────────────────────────────────────────────────────
# Enrichment — Easy Apply + company info + experience filter
# ─────────────────────────────────────────────────────────────────
def enrich_jobs(jobs, cfg):
    min_yrs   = cfg.get("min_years_experience", 0)
    filter_on = min_yrs > 0
    passed, dropped = [], 0

    log(f"  Enriching {len(jobs)} new jobs...")
    if not filter_on:
        log("  Experience filter: DISABLED")

    for i, job in enumerate(jobs, 1):
        url = job.get("LinkedIn URL", "")
        log(f"    [{i}/{len(jobs)}] {job.get('Company','?')} — {job.get('Job Title','?')}")

        # ── Detail page: Easy Apply + closed check + recruiter ─────
        detail = fetch_job_detail(url)
        time.sleep(1)

        job["Easy Apply"]        = "Yes" if detail["easy_apply"] else "No"
        job["Recruiter Profile"] = detail["recruiter"] if detail["easy_apply"] else ""

        # Still Accepting? column
        is_closed = detail.get("is_closed", False)
        job["Still Accepting?"] = "No" if is_closed else "Yes"

        # Closed jobs filter
        if cfg.get("exclude_closed_jobs", True) and is_closed:
            dropped += 1
            log(f"      ✗ Dropped — job closed: {job.get('Job Title','?')} @ {job.get('Company','?')}")
            continue

        # ── Keyword scoring against full description ──────────────
        desc_text     = detail.get("description", "")
        full_text     = (job.get("Job Title","") + " " +
                         job.get("Company","") + " " +
                         job.get("Location","") + " " +
                         desc_text).lower()
        title_text    = job.get("Job Title","").lower()

        primary_kws   = [k.strip().lower() for k in cfg.get("primary_must_have", []) if k.strip()]
        secondary_kws = [k.strip().lower() for k in cfg.get("secondary_must_have", []) if k.strip()]
        preferred_kws = [k.lower() for k in cfg.get("preferred_keywords", [])]

        keywords_configured = bool(primary_kws or secondary_kws)

        # Primary match: YES only if ALL primary keywords are present
        if primary_kws:
            primary_hit = all(kw in full_text for kw in primary_kws)
            job["Primary Keywords Match"] = "Yes" if primary_hit else "No"
        else:
            job["Primary Keywords Match"] = "NA"
            primary_hit = None   # not configured

        # Secondary match: YES if at least ONE secondary keyword is present
        if secondary_kws:
            secondary_hit = any(kw in full_text for kw in secondary_kws)
            job["Secondary Keywords Match"] = "Yes" if secondary_hit else "No"
        else:
            job["Secondary Keywords Match"] = "NA"
            secondary_hit = None   # not configured

        # ── Score (0–100) ─────────────────────────────────────────
        # Breakdown:
        #   40 pts — title relevance (preferred keywords in title)
        #   35 pts — primary keywords (all present = full 35 pts,
        #             partial = proportional, not configured = 0)
        #   25 pts — secondary keywords (any present = 25 pts)
        # If NO keywords configured at all → "NA"
        if not keywords_configured:
            job["Score"] = "NA"
        else:
            # Title relevance (max 40)
            title_hits   = sum(1 for p in preferred_kws if p in title_text)
            title_max    = max(len(preferred_kws), 1)
            title_score  = round((title_hits / title_max) * 40)

            # Primary score (max 35) — proportional to how many matched
            if primary_kws:
                p_matched   = sum(1 for kw in primary_kws if kw in full_text)
                primary_score = round((p_matched / len(primary_kws)) * 35)
            else:
                primary_score = 0

            # Secondary score (max 25) — binary: any hit = full 25
            if secondary_kws:
                secondary_score = 25 if secondary_hit else 0
            else:
                secondary_score = 0

            job["Score"] = min(title_score + primary_score + secondary_score, 100)


        # ── Company page: size + type ──────────────────────────────
        company_url  = get_company_page_url(url)
        company_data = scrape_company_bio(company_url)
        job["Company Size"] = company_data["company_size"]
        job["Company Type"] = company_data["company_type"]
        time.sleep(1.5)

        job["Applied?"] = ""

        # ── Experience filter ──────────────────────────────────────
        if filter_on:
            years = extract_min_years(detail.get("description", ""))
            if years is None:
                job["Exp. Required"] = "Not specified"
                passed.append(job)
            elif years >= min_yrs:
                job["Exp. Required"] = f"{years}+ yrs"
                passed.append(job)
            else:
                dropped += 1
                log(f"      ✗ Dropped ({years} yrs required)")
        else:
            job["Exp. Required"] = "Filter off"
            passed.append(job)

    if filter_on:
        log(f"  Experience filter: {len(passed)} kept, {dropped} dropped")
    return passed


# ─────────────────────────────────────────────────────────────────
# Excel — save consolidated file
# ─────────────────────────────────────────────────────────────────
def apply_styles_and_dropdown(ws, total_rows):
    from openpyxl.styles import PatternFill, Font
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill as PFill

    col_widths = {
        "A": 14,  # Applied?
        "B": 34,  # Job Title
        "C": 22,  # Company
        "D": 22,  # Company Size
        "E": 28,  # Company Type
        "F": 16,  # Level
        "G": 18,  # Location
        "H": 12,  # Posted Date
        "I": 12,  # Easy Apply
        "J": 42,  # Recruiter Profile
        "K": 10,  # Score
        "L": 22,  # Primary Keywords Match
        "M": 24,  # Secondary Keywords Match
        "N": 14,  # Work Mode
        "O": 16,  # Still Accepting?
        "P": 18,  # Date Searched
        "Q": 14,  # Exp. Required
        "R": 65,  # LinkedIn URL
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "B2"

    score_col_idx   = COL_ORDER.index("Score")
    url_col_idx     = COL_ORDER.index("LinkedIn URL")
    rec_col_idx     = COL_ORDER.index("Recruiter Profile")
    applied_col_idx = COL_ORDER.index("Applied?")
    title_col_idx   = COL_ORDER.index("Job Title")

    # Score-based row colours (when Applied? is blank)
    # 80-100 → green, 50-79 → blue, 0-49 → no fill, NA → no fill
    fill_high = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fill_mid  = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        applied_val = str(row[applied_col_idx].value or "").strip()
        score_val   = row[score_col_idx].value
        if not applied_val:
            try:
                s = int(float(str(score_val)))
                if s >= 80:
                    for cell in row: cell.fill = fill_high
                elif s >= 50:
                    for cell in row: cell.fill = fill_mid
            except (TypeError, ValueError):
                pass   # NA or blank — no fill
        # LinkedIn URL hyperlink
        url_cell = row[url_col_idx]
        if url_cell.value and str(url_cell.value).startswith("http"):
            url_cell.hyperlink = url_cell.value
            url_cell.font = Font(color="0563C1", underline="single")
        # Recruiter hyperlink
        rec_cell = row[rec_col_idx]
        if rec_cell.value and str(rec_cell.value).startswith("http"):
            rec_cell.hyperlink = rec_cell.value
            rec_cell.font = Font(color="0563C1", underline="single")
        # Job Title hyperlink
        title_cell  = row[title_col_idx]
        job_url_val = row[url_col_idx].value
        if title_cell.value and job_url_val and str(job_url_val).startswith("http"):
            title_cell.hyperlink = str(job_url_val)
            title_cell.font = Font(color="0563C1", underline="single", bold=True)

    # Dropdown
    options_str = ",".join(APPLIED_OPTIONS[1:])
    dv = DataValidation(type="list", formula1=f'"{options_str}"',
                        allow_blank=True, showDropDown=False, showErrorMessage=False)
    dv.sqref = f"A2:A{max(total_rows + 1, 5000)}"
    ws.add_data_validation(dv)

    # Conditional formatting on Applied? column
    cf_range = f"A2:A{max(total_rows + 1, 5000)}"
    ws.conditional_formatting.add(cf_range, CellIsRule(
        operator="equal", formula=['"Relevant"'],
        fill=PFill(start_color="BBFBD0", end_color="BBFBD0", fill_type="solid")))
    ws.conditional_formatting.add(cf_range, CellIsRule(
        operator="equal", formula=['"Applied"'],
        fill=PFill(start_color="BFDBFE", end_color="BFDBFE", fill_type="solid")))
    ws.conditional_formatting.add(cf_range, CellIsRule(
        operator="equal", formula=['"Irrelevant - Criteria mismatch"'],
        fill=PFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid")))

    # Keyword match column colouring
    kw_cols = ["L", "M"]   # Primary Keywords Match, Secondary Keywords Match
    for col_letter in kw_cols:
        kw_range = f"{col_letter}2:{col_letter}{max(total_rows + 1, 5000)}"
        ws.conditional_formatting.add(kw_range, CellIsRule(
            operator="equal", formula=['"Yes"'],
            fill=PFill(start_color="BBFBD0", end_color="BBFBD0", fill_type="solid")))
        ws.conditional_formatting.add(kw_range, CellIsRule(
            operator="equal", formula=['"No"'],
            fill=PFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")))


def is_file_locked(filepath):
    if not os.path.exists(filepath):
        return False
    try:
        with open(filepath, "a"):
            return False
    except IOError:
        return True


def save_excel(jobs, cfg):
    # Score column used for sorting — numeric desc, NA last
    filepath  = os.path.join(cfg["output_folder"], cfg["output_filename"])

    new_df = pd.DataFrame(jobs) if jobs else pd.DataFrame(columns=COL_ORDER)
    for col in COL_ORDER:
        if col not in new_df.columns:
            new_df[col] = ""
    new_df = new_df[COL_ORDER]

    existing_df = pd.DataFrame(columns=COL_ORDER)
    applied_map = {}

    if os.path.exists(filepath):
        try:
            existing_df = pd.read_excel(filepath, sheet_name="LinkedIn Jobs", dtype=str)
            existing_df.fillna("", inplace=True)
            for col in COL_ORDER:
                if col not in existing_df.columns:
                    existing_df[col] = ""
            existing_df = existing_df[COL_ORDER]
            for _, row in existing_df.iterrows():
                val = str(row.get("Applied?", "")).strip()
                if not val:
                    continue
                url = str(row.get("LinkedIn URL", "")).strip()
                key = (str(row.get("Job Title","")) + str(row.get("Company",""))).lower().strip()
                if url: applied_map[url] = val
                if key: applied_map[key] = val
            log(f"  Loaded {len(existing_df)} existing rows ({len(applied_map)} with Applied? set)")
        except Exception as e:
            log(f"  Could not read existing file ({e}) — creating fresh.")
            existing_df = pd.DataFrame(columns=COL_ORDER)

    if not existing_df.empty and not new_df.empty:
        existing_urls = set(existing_df["LinkedIn URL"].astype(str))
        existing_keys = set((str(r["Job Title"]) + str(r["Company"])).lower().strip()
                            for _, r in existing_df.iterrows())
        def is_new(row):
            url = str(row.get("LinkedIn URL", ""))
            key = (str(row.get("Job Title","")) + str(row.get("Company",""))).lower().strip()
            return url not in existing_urls and key not in existing_keys
        new_df = new_df[new_df.apply(is_new, axis=1)]

    if new_df.empty and not existing_df.empty:
        log("  No new rows to add — re-saving to apply style updates.")
        combined_df = existing_df
    elif existing_df.empty:
        combined_df = new_df
        log(f"  Creating new file with {len(combined_df)} jobs")
    else:
        combined_df = pd.concat([new_df, existing_df], ignore_index=True)
        log(f"  Adding {len(new_df)} new row(s) — total: {len(combined_df)}")

    # Restore Applied? values
    def restore_applied(row):
        url = str(row.get("LinkedIn URL", "")).strip()
        key = (str(row.get("Job Title","")) + str(row.get("Company",""))).lower().strip()
        return applied_map.get(url) or applied_map.get(key) or row.get("Applied?", "")
    combined_df["Applied?"] = combined_df.apply(restore_applied, axis=1)

    # Sort: newest first, then match quality
    # Sort: latest Date Searched first, then Score descending (NA sorts last)
    def score_sort_key(val):
        try:    return -int(float(str(val)))   # negative so higher score = earlier row
        except: return 1                        # "NA" or blank sorts to end
    combined_df["_score_key"] = combined_df["Score"].apply(score_sort_key)
    combined_df["_dt"]        = pd.to_datetime(combined_df["Date Searched"], errors="coerce")
    combined_df = combined_df.sort_values(["_dt", "_score_key"], ascending=[False, True])
    combined_df = combined_df.drop(columns=["_score_key", "_dt"])
    combined_df["Date Searched"] = pd.to_datetime(
        combined_df["Date Searched"], errors="coerce"
    ).dt.strftime("%Y-%m-%d %H:%M")
    combined_df.fillna("", inplace=True)

    write_path = filepath
    if is_file_locked(filepath):
        ts = datetime.now().strftime("%H%M%S")
        write_path = filepath.replace(".xlsx", f"_backup_{ts}.xlsx")
        log(f"  ⚠ File open in Excel — saving to: {os.path.basename(write_path)}")

    try:
        with pd.ExcelWriter(write_path, engine="openpyxl") as writer:
            combined_df.to_excel(writer, index=False, sheet_name="LinkedIn Jobs")
            apply_styles_and_dropdown(writer.sheets["LinkedIn Jobs"], len(combined_df))
        log(f"  Saved → {write_path}  ({len(combined_df)} total rows)")
        return write_path
    except PermissionError:
        log("  ✗ Close the Excel file and try again.")
        return None
    except Exception as e:
        log(f"  ✗ Save error: {e}")
        return None


# ─────────────────────────────────────────────────────────────────
# Email
# ─────────────────────────────────────────────────────────────────
def send_email(filepath, count, cfg):
    if not cfg["email_from"] or not cfg["email_to"]:
        return
    try:
        msg = MIMEMultipart()
        msg["From"]    = cfg["email_from"]
        msg["To"]      = cfg["email_to"]
        msg["Subject"] = f"LinkedIn TA Jobs — {count} new ({datetime.now().strftime('%d %b %Y')})"
        body = (
            f"Role      : {cfg['job_title']}\n"
            f"Locations : {', '.join(as_list(cfg['location']))}\n"
            f"Work modes: {', '.join(as_list(cfg['work_mode']))}\n"
            f"Posted    : last {cfg['posted_within_days']} day(s)\n"
            f"Min exp   : {cfg.get('min_years_experience', 0)}+ years\n"
            f"New jobs  : {count}\n\n"
            "See attached consolidated Excel.\n\n— LinkedIn Job Scheduler"
        )
        msg.attach(MIMEText(body, "plain"))
        if filepath and os.path.exists(filepath):
            with open(filepath, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition",
                                f'attachment; filename="{os.path.basename(filepath)}"')
                msg.attach(part)
        with smtplib.SMTP(cfg["email_smtp"], cfg["email_port"]) as s:
            s.starttls()
            s.login(cfg["email_from"], cfg["email_password"])
            s.send_message(msg)
        log("  Email sent.")
    except Exception as e:
        log(f"  Email failed: {e}")


# ─────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────
def run_daily_search():
    log("=" * 60)
    log("LinkedIn Job Scheduler — run starting")
    log(f"  Role      : {CONFIG['job_title']}")
    log(f"  Locations : {as_list(CONFIG['location'])}")
    log(f"  Work modes: {as_list(CONFIG['work_mode'])}")
    log(f"  Exp levels: {as_list(CONFIG['experience_level'])}")
    log(f"  Posted    : last {CONFIG['posted_within_days']} day(s)")
    log(f"  Min exp   : {CONFIG.get('min_years_experience', 0)}+ yrs")
    log("")

    log("Step 1/4 — Fetching jobs from LinkedIn...")
    jobs = fetch_all_combinations(CONFIG)
    log(f"  Unique jobs fetched: {len(jobs)}")

    log("Step 2/4 — Deduplication (before enriching)...")
    jobs = filter_new_only(jobs, CONFIG)
    if not jobs:
        log("  No new jobs found — nothing to do.")
        log("Done.\n")
        return

    log("Step 3/4 — Enriching new jobs (Easy Apply + company info)...")
    jobs = enrich_jobs(jobs, CONFIG)

    log("Step 4/4 — Saving consolidated Excel...")
    filepath = save_excel(jobs, CONFIG)
    send_email(filepath, len(jobs), CONFIG)
    log("Done.\n")


if __name__ == "__main__":
    log(f"Scheduler started — daily run at {CONFIG['schedule_time']}.")
    log("Press Ctrl+C to stop.\n")
    run_daily_search()
    schedule.every().day.at(CONFIG["schedule_time"]).do(run_daily_search)
    while True:
        schedule.run_pending()
        time.sleep(30)
